import build_conditions as bc, routeb as rb, scen_codec as sc, unicodedata, re
from openpyxl import load_workbook

def parse_endings(path):
    ws=load_workbook(path, read_only=True)['Sheet1']
    rows=[(r[0] if r and r[0] is not None else '') for r in ws.iter_rows(values_only=True)]
    blocks=[]; pages=[]; lines=[]
    for raw in rows:
        t=str(raw).strip()
        if t=='': continue
        endb=t.endswith('<>'); brk='<clsr>' in t
        line=t.replace('<clsr>','').replace('<>','').strip()
        if line: lines.append(line)
        if brk: pages.append(lines); lines=[]
        if endb:
            if lines: pages.append(lines); lines=[]
            blocks.append(pages); pages=[]
    if lines: pages.append(lines)
    if pages: blocks.append(pages)
    return blocks

def norm(s):
    s=unicodedata.normalize('NFKC',s)
    return (s.replace('\u2019',"'").replace('\u2018',"'")
             .replace('\u201c','"').replace('\u201d','"')
             .replace('\u2014','-').replace('\u2013','-').replace('\u2026','...'))

def wrap(words, width=28):
    lines=[]; cur=''
    for w in words:
        if not cur: cur=w
        elif len(cur)+1+len(w)<=width: cur+=' '+w
        else: lines.append(cur); cur=w
    if cur: lines.append(cur)
    return lines

def encode_epilogue(pages, width=28, maxlines=3):
    # pages: list of pages, each a list of authored lines. Re-wrap each page, paginate to maxlines.
    boxes=[]
    for pg in pages:
        text=norm(' '.join(pg))
        wl=wrap(text.split(), width)
        for k in range(0,len(wl),maxlines):
            boxes.append(wl[k:k+maxlines])
    # encode: lines -> {08} join; boxes -> {06}{07} join
    out=bytearray()
    for bi,box in enumerate(boxes):
        if bi>0: out+=bytes([0x06,0x07])
        for li,ln in enumerate(box):
            if li>0: out+=bytes([0x08])
            out+=bc.enc_text_run(ln)
    return bytes(out)

if __name__=='__main__':
    blocks=parse_endings('/mnt/user-data/uploads/langendings.xlsx')
    before=len(bc.new_glyphs)
    enc=[encode_epilogue(p) for p in blocks]
    print('epilogues:',len(enc))
    print('new glyphs needed by endings:', len(bc.new_glyphs)-before)
    print('blank slots available total:', len(bc.free))
    tot=sum(len(x) for x in enc)+len(enc)  # +nulls
    print('encoded endings bytes ~', tot, '(JP entry9 was 3784)')
    # measure SCEN size if applied (conditions + endings)
    m=bc.model
    for b in m['blocks']:
        cnt,offs,ents=sc.parse_section2(b['sections'][2])
        s6=ents[6].split(b'\x00'); ents[6]=b'\x00'.join(bc.C(*bc.TR[s.decode('shift_jis','replace')]) if s.decode('shift_jis','replace') in bc.TR else s for s in s6)
        s4=ents[4].split(b'\x00'); ents[4]=b'\x00'.join(bc.C(*bc.E4TR[s.decode('shift_jis','replace')]) if s.decode('shift_jis','replace') in bc.E4TR else s for s in s4)
        b['sections'][2]=sc.build_section2(ents)
    c,off,e=sc.parse_section2(m['blocks'][17]['sections'][2])
    e[9]=b'\x00'.join(enc)
    m['blocks'][17]['sections'][2]=sc.build_section2(e)
    out=sc.serialize(m)
    print('SCEN with conditions+endings:', hex(len(out)), '=', (len(out)+2047)//2048, 'sectors (budget 364)')
